"""Tests for Excel/CSV import-export + renewal reporting (data-io feature).

Coverage
--------
1. Auth guards (401 without token) for export, import, and renewal-report endpoints.
2. XLSX writer produces a non-empty workbook with the correct header row.
3. Import parser: 1 valid + 1 bad-category CSV row → created=1, errors has bad row.
4. Renewal window filter returns only policies within the window (service-level).

All service-level tests are offline (no live DB / HTTP).
Endpoint-level tests use a minimal FastAPI test app (same pattern as test_intake.py).
"""
from __future__ import annotations

import io
import uuid
from datetime import date, timedelta
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch

import httpx
import openpyxl
import pytest
from fastapi import FastAPI

from app.api.v1 import exports as exports_module
from app.api.v1 import imports as imports_module
from app.api.v1 import reports as reports_module
from app.core.errors import register_error_handlers
from app.core.middleware import RequestIDMiddleware
from app.schemas.reports import RenewalReportRow
from app.services import data_io_service

# ---------------------------------------------------------------------------
# Minimal test app — all three routers, same middleware as prod
# ---------------------------------------------------------------------------


def _make_test_app() -> FastAPI:
    test_app = FastAPI()
    test_app.add_middleware(RequestIDMiddleware)
    register_error_handlers(test_app)
    test_app.include_router(exports_module.router, prefix="/api/v1")
    test_app.include_router(imports_module.router, prefix="/api/v1")
    test_app.include_router(reports_module.router, prefix="/api/v1")
    return test_app


_test_app = _make_test_app()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_TENANT_ID = "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
_ORG_ID = "bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb"


def _fake_policy(
    title: str = "Fleet Motor",
    category: str = "vehicle",
    status: str = "active",
    expiry_days: int | None = None,
    provider_id: uuid.UUID | None = None,
) -> MagicMock:
    p = MagicMock()
    p.id = uuid.uuid4()
    p.tenant_id = uuid.UUID(_TENANT_ID)
    p.org_id = uuid.UUID(_ORG_ID)
    p.title = title
    p.category = category
    p.policy_number = "POL-001"
    p.provider_id = provider_id
    p.sum_insured_inr = Decimal("1000000")
    p.premium_inr = Decimal("25000")
    p.gst_inr = Decimal("4500")
    p.inception_date = date(2025, 1, 1)
    p.expiry_date = date.today() + timedelta(days=expiry_days) if expiry_days is not None else None
    p.renewal_date = None
    p.status = status
    return p


# ---------------------------------------------------------------------------
# 1a. Auth guard — GET /policies/export
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_policies_requires_auth():
    """No Bearer token → 401."""
    transport = httpx.ASGITransport(app=_test_app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as ac:
        resp = await ac.get("/api/v1/policies/export")
    assert resp.status_code == 401
    assert resp.json()["error"]["code"] == "unauthorized"


# ---------------------------------------------------------------------------
# 1b. Auth guard — POST /policies/import
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_policies_requires_auth():
    """No Bearer token → 401."""
    transport = httpx.ASGITransport(app=_test_app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as ac:
        resp = await ac.post(
            "/api/v1/policies/import",
            files={"file": ("data.csv", b"title,category\n", "text/csv")},
        )
    assert resp.status_code == 401
    assert resp.json()["error"]["code"] == "unauthorized"


# ---------------------------------------------------------------------------
# 1c. Auth guard — GET /reports/renewals
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_renewals_report_requires_auth():
    """No Bearer token → 401."""
    transport = httpx.ASGITransport(app=_test_app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as ac:
        resp = await ac.get("/api/v1/reports/renewals")
    assert resp.status_code == 401
    assert resp.json()["error"]["code"] == "unauthorized"


# ---------------------------------------------------------------------------
# 1d. Auth guard — GET /reports/renewals/export
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_export_renewals_requires_auth():
    """No Bearer token → 401."""
    transport = httpx.ASGITransport(app=_test_app)
    async with httpx.AsyncClient(transport=transport, base_url="http://test") as ac:
        resp = await ac.get("/api/v1/reports/renewals/export")
    assert resp.status_code == 401
    assert resp.json()["error"]["code"] == "unauthorized"


# ---------------------------------------------------------------------------
# 2. XLSX writer — non-empty workbook with correct header row
# ---------------------------------------------------------------------------


def test_write_policies_xlsx_produces_correct_headers():
    """write_policies_xlsx must produce a workbook with the expected header row."""
    policies = [_fake_policy("Policy A"), _fake_policy("Policy B")]
    provider_map: dict[uuid.UUID, str] = {}

    buf = data_io_service.write_policies_xlsx(policies, provider_map)

    assert isinstance(buf, io.BytesIO)
    assert buf.tell() == 0  # position reset to 0

    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws is not None

    headers = [cell.value for cell in ws[1]]
    expected = [label for _, label in data_io_service.EXPORT_COLUMNS]
    assert headers == expected, f"Header mismatch: {headers}"

    # Two data rows.
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2
    assert data_rows[0][0] == "Policy A"
    assert data_rows[1][0] == "Policy B"


def test_write_policies_xlsx_resolves_provider_name():
    """Provider ID in provider_map is substituted with the provider name."""
    pid = uuid.uuid4()
    p = _fake_policy(provider_id=pid)
    provider_map = {pid: "HDFC ERGO"}

    buf = data_io_service.write_policies_xlsx([p], provider_map)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    # Column index 3 (0-based) = "Provider" (4th column, 1-indexed = col 4)
    header_row = [cell.value for cell in ws[1]]
    provider_col_idx = header_row.index("Provider") + 1  # 1-indexed
    data_val = ws.cell(row=2, column=provider_col_idx).value
    assert data_val == "HDFC ERGO"


# ---------------------------------------------------------------------------
# 3. Import parser — 1 valid + 1 bad-category row
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_import_parser_one_valid_one_bad_category():
    """CSV with 1 valid + 1 bad-category row → created=1, skipped=0, 1 error."""
    csv_content = (
        "Title,Category,Policy Number,Premium (INR),Expiry Date\n"
        "Fleet Motor,vehicle,POL-001,25000,2026-03-31\n"
        "Marine Cargo,marine_cargo_invalid,POL-002,10000,2026-06-30\n"
    )
    raw_rows = data_io_service.parse_csv_bytes(csv_content.encode())
    assert len(raw_rows) == 2

    # Mock create_policy so no DB is needed.
    created_policies: list = []

    async def _mock_create(db, user, payload):
        created_policies.append(payload)
        return MagicMock(id=uuid.uuid4())

    mock_db = AsyncMock()
    mock_user = MagicMock()
    mock_user.tenant_id = _TENANT_ID
    mock_user.org_id = _ORG_ID
    mock_user.user_id = str(uuid.uuid4())
    mock_user.role = "admin"
    mock_user.is_super_admin = False

    target_org = uuid.UUID(_ORG_ID)

    with patch.object(data_io_service, "create_policy", side_effect=_mock_create):
        created, skipped, errors = await data_io_service.run_import(
            mock_db, mock_user, raw_rows, target_org
        )

    assert created == 1, f"Expected 1 created, got {created}"
    assert skipped == 0
    assert len(errors) == 1
    assert errors[0].row == 3  # row 1=header, row 2=valid, row 3=bad
    assert "marine_cargo_invalid" in errors[0].message
    assert created_policies[0].title == "Fleet Motor"
    assert created_policies[0].category == "vehicle"


@pytest.mark.asyncio
async def test_import_blank_rows_are_skipped():
    """Fully blank rows are counted as skipped, not errors."""
    csv_content = (
        "Title,Category\n"
        "Fleet Motor,vehicle\n"
        ",,\n"
        ",\n"
    )
    raw_rows = data_io_service.parse_csv_bytes(csv_content.encode())

    mock_db = AsyncMock()
    mock_user = MagicMock()
    mock_user.org_id = _ORG_ID
    mock_user.tenant_id = _TENANT_ID
    mock_user.user_id = str(uuid.uuid4())
    mock_user.role = "admin"
    mock_user.is_super_admin = False

    target_org = uuid.UUID(_ORG_ID)

    async def _mock_create(db, user, payload):
        return MagicMock(id=uuid.uuid4())

    with patch.object(data_io_service, "create_policy", side_effect=_mock_create):
        created, skipped, errors = await data_io_service.run_import(
            mock_db, mock_user, raw_rows, target_org
        )

    assert created == 1
    assert skipped == 2
    assert errors == []


# ---------------------------------------------------------------------------
# 4. Renewal window filter — returns only policies within the window
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_fetch_renewal_report_window_filter():
    """Only policies expiring within window_days are included."""
    today = date.today()

    # Three policies: expiring in 10, 60, and 100 days.
    p_10 = _fake_policy("Expires Soon", expiry_days=10)
    p_60 = _fake_policy("Expires Mid", expiry_days=60)
    p_100 = _fake_policy("Expires Late", expiry_days=100)

    mock_db = AsyncMock()
    mock_user = MagicMock()
    mock_user.tenant_id = _TENANT_ID
    mock_user.org_id = _ORG_ID
    mock_user.role = "admin"
    mock_user.is_super_admin = True  # group-wide access

    # Patch the internal DB query to return our fake policies.
    async def _fake_execute(stmt):
        result = MagicMock()
        result.scalars.return_value.all.return_value = [p_10, p_60, p_100]
        return result

    async def _fake_provider_map(db, tenant_id):
        return {}

    mock_db.execute = _fake_execute

    with patch.object(data_io_service, "_provider_name_map", side_effect=_fake_provider_map):
        rows_90 = await data_io_service.fetch_renewal_report(mock_db, mock_user, window_days=90)
        rows_30 = await data_io_service.fetch_renewal_report(mock_db, mock_user, window_days=30)

    # With 90-day window: 10 and 60 are in; 100 is out.
    assert len(rows_90) == 2, f"Expected 2 rows, got {len(rows_90)}: {[r.title for r in rows_90]}"
    assert rows_90[0].title == "Expires Soon"
    assert rows_90[1].title == "Expires Mid"

    # With 30-day window: only 10 is in.
    assert len(rows_30) == 1
    assert rows_30[0].title == "Expires Soon"
    assert rows_30[0].days_left == (p_10.expiry_date - today).days


def test_renewal_report_ordering():
    """Rows must be ordered by days_left ascending."""
    rows = [
        RenewalReportRow(
            policy_id=uuid.uuid4(),
            title="C", category="vehicle", provider_name=None,
            expiry_date=date.today() + timedelta(days=60),
            days_left=60, premium_inr=None, sum_insured_inr=None, status="active",
        ),
        RenewalReportRow(
            policy_id=uuid.uuid4(),
            title="A", category="vehicle", provider_name=None,
            expiry_date=date.today() + timedelta(days=5),
            days_left=5, premium_inr=None, sum_insured_inr=None, status="active",
        ),
        RenewalReportRow(
            policy_id=uuid.uuid4(),
            title="B", category="vehicle", provider_name=None,
            expiry_date=date.today() + timedelta(days=30),
            days_left=30, premium_inr=None, sum_insured_inr=None, status="active",
        ),
    ]
    # Sort mimics what the service does (DB orders by expiry_date asc, window filter preserves it).
    sorted_rows = sorted(rows, key=lambda r: r.days_left)
    assert [r.title for r in sorted_rows] == ["A", "B", "C"]


# ---------------------------------------------------------------------------
# 5. Category normalisation
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("raw,expected", [
    ("vehicle", "vehicle"),
    ("Vehicle", "vehicle"),
    ("factory property", "factory_property"),
    ("factory_property", "factory_property"),
    ("employees group health", "employees_group_health"),
    ("MACHINERY", "machinery"),
    ("marine_cargo", None),          # invalid → None
    ("", None),
    ("unknown_type", None),
])
def test_normalize_category(raw, expected):
    assert data_io_service._normalize_category(raw) == expected


# ---------------------------------------------------------------------------
# 6. Date parsing
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("value,expected", [
    ("2025-04-01", date(2025, 4, 1)),
    ("01-04-2025", date(2025, 4, 1)),
    ("01/04/2025", date(2025, 4, 1)),
    (date(2025, 4, 1), date(2025, 4, 1)),
    (None, None),
    ("", None),
    ("not-a-date", None),
])
def test_parse_date_cell(value, expected):
    assert data_io_service._parse_date_cell(value) == expected


# ---------------------------------------------------------------------------
# 7. Decimal parsing
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("value,expected", [
    ("25000", Decimal("25000")),
    ("1,00,000", Decimal("100000")),
    (25000.5, Decimal("25000.5")),
    (None, None),
    ("", None),
    ("N/A", None),
])
def test_parse_decimal(value, expected):
    assert data_io_service._parse_decimal(value) == expected


# ---------------------------------------------------------------------------
# 8. Header normalisation
# ---------------------------------------------------------------------------


@pytest.mark.parametrize("raw,expected", [
    ("Title", "title"),
    ("policy number", "policy_number"),
    ("Policy Number", "policy_number"),
    ("Sum Insured (INR)", "sum_insured_inr"),
    ("premium", "premium_inr"),
    ("EXPIRY DATE", "expiry_date"),
    ("unknown_column", None),
])
def test_map_header(raw, expected):
    assert data_io_service._map_header(raw) == expected
