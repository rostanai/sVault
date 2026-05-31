"""Data I/O service — Excel/CSV export + import + renewal reporting.

All query helpers here are tenant/org-scoped.  They reuse the same
_accessible_org_filter logic from policy_service.  No FastAPI imports.
"""
from __future__ import annotations

import csv
import io
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.errors import AppError, ErrorCode
from app.core.security import CurrentUser
from app.db.models import Policy, Provider
from app.schemas.policy import PolicyCreate
from app.schemas.reports import ImportRowError, RenewalReportRow
from app.services.org_service import is_group_wide
from app.services.policy_service import create_policy

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MAX_IMPORT_ROWS = 1000
MAX_FILE_BYTES = 5 * 1024 * 1024  # 5 MB

# Export column order and header labels.
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("title", "Title"),
    ("category", "Category"),
    ("policy_number", "Policy Number"),
    ("provider_name", "Provider"),
    ("sum_insured_inr", "Sum Insured (INR)"),
    ("premium_inr", "Premium (INR)"),
    ("gst_inr", "GST (INR)"),
    ("inception_date", "Inception Date"),
    ("expiry_date", "Expiry Date"),
    ("renewal_date", "Renewal Date"),
    ("status", "Status"),
]

# Canonical import aliases (lower-cased, stripped) → PolicyCreate field.
_IMPORT_ALIASES: dict[str, str] = {
    "title": "title",
    "category": "category",
    "policy number": "policy_number",
    "policy_number": "policy_number",
    "policynumber": "policy_number",
    "provider": "provider_name",
    "provider name": "provider_name",
    "provider_name": "provider_name",
    "sum insured (inr)": "sum_insured_inr",
    "sum_insured_inr": "sum_insured_inr",
    "sum insured": "sum_insured_inr",
    "premium (inr)": "premium_inr",
    "premium_inr": "premium_inr",
    "premium": "premium_inr",
    "gst (inr)": "gst_inr",
    "gst_inr": "gst_inr",
    "gst": "gst_inr",
    "inception date": "inception_date",
    "inception_date": "inception_date",
    "inceptiondate": "inception_date",
    "expiry date": "expiry_date",
    "expiry_date": "expiry_date",
    "expirydate": "expiry_date",
    "renewal date": "renewal_date",
    "renewal_date": "renewal_date",
    "renewaldate": "renewal_date",
    "status": "status",
}

# Valid PolicyCategory values.
_VALID_CATEGORIES: set[str] = {
    "vehicle", "machinery", "plant", "factory_property",
    "employees_group_health", "key_person",
    "stock_raw_material", "stock_finished_goods", "other",
}


# ---------------------------------------------------------------------------
# Scoping helper (mirrors policy_service)
# ---------------------------------------------------------------------------

def _accessible_org_filter(user: CurrentUser) -> uuid.UUID | None:
    if user.is_super_admin or is_group_wide(user.role):
        return None
    return uuid.UUID(user.org_id) if user.org_id else None


# ---------------------------------------------------------------------------
# Provider name lookup — single query, returns {provider_id: name}
# ---------------------------------------------------------------------------

async def _provider_name_map(
    db: AsyncSession, tenant_id: uuid.UUID
) -> dict[uuid.UUID, str]:
    stmt = select(Provider.id, Provider.name).where(Provider.tenant_id == tenant_id)
    rows = (await db.execute(stmt)).all()
    return {r.id: r.name for r in rows}


# ---------------------------------------------------------------------------
# Query helpers
# ---------------------------------------------------------------------------

async def _fetch_policies_for_export(
    db: AsyncSession, user: CurrentUser
) -> tuple[list[Policy], dict[uuid.UUID, str]]:
    """Return all accessible policies + a provider-name map (no N+1)."""
    tid = uuid.UUID(user.tenant_id)
    stmt = select(Policy).where(Policy.tenant_id == tid)
    org = _accessible_org_filter(user)
    if org is not None:
        stmt = stmt.where(Policy.org_id == org)
    stmt = stmt.order_by(Policy.expiry_date.asc().nullslast())
    policies = list((await db.execute(stmt)).scalars().all())
    provider_map = await _provider_name_map(db, tid)
    return policies, provider_map


async def fetch_renewal_report(
    db: AsyncSession, user: CurrentUser, window_days: int = 90
) -> list[RenewalReportRow]:
    """Return policies whose expiry_date is within the next `window_days` days."""
    tid = uuid.UUID(user.tenant_id)
    today = date.today()

    stmt = select(Policy).where(
        Policy.tenant_id == tid,
        Policy.expiry_date.isnot(None),
    )
    org = _accessible_org_filter(user)
    if org is not None:
        stmt = stmt.where(Policy.org_id == org)
    stmt = stmt.order_by(Policy.expiry_date.asc())

    policies = list((await db.execute(stmt)).scalars().all())
    provider_map = await _provider_name_map(db, tid)

    rows: list[RenewalReportRow] = []
    for p in policies:
        if p.expiry_date is None:
            continue
        days_left = (p.expiry_date - today).days
        if days_left < 0 or days_left > window_days:
            continue
        rows.append(
            RenewalReportRow(
                policy_id=p.id,
                title=p.title,
                category=p.category,
                provider_name=provider_map.get(p.provider_id) if p.provider_id else None,
                expiry_date=p.expiry_date,
                days_left=days_left,
                premium_inr=p.premium_inr,
                sum_insured_inr=p.sum_insured_inr,
                status=p.status,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Export writers
# ---------------------------------------------------------------------------

def _policy_row(p: Any, provider_map: dict[uuid.UUID, str]) -> dict[str, Any]:
    """Convert a Policy ORM object (or duck-typed dict-like) to a flat export dict."""
    provider_id = getattr(p, "provider_id", None)
    provider_name = provider_map.get(provider_id, str(provider_id)) if provider_id else ""
    return {
        "title": getattr(p, "title", "") or "",
        "category": getattr(p, "category", "") or "",
        "policy_number": getattr(p, "policy_number", "") or "",
        "provider_name": provider_name,
        "sum_insured_inr": _fmt_decimal(getattr(p, "sum_insured_inr", None)),
        "premium_inr": _fmt_decimal(getattr(p, "premium_inr", None)),
        "gst_inr": _fmt_decimal(getattr(p, "gst_inr", None)),
        "inception_date": _fmt_date(getattr(p, "inception_date", None)),
        "expiry_date": _fmt_date(getattr(p, "expiry_date", None)),
        "renewal_date": _fmt_date(getattr(p, "renewal_date", None)),
        "status": getattr(p, "status", "") or "",
    }


def _renewal_row(r: RenewalReportRow) -> dict[str, Any]:
    return {
        "policy_id": str(r.policy_id),
        "title": r.title,
        "category": r.category,
        "provider_name": r.provider_name or "",
        "expiry_date": _fmt_date(r.expiry_date),
        "days_left": str(r.days_left) if r.days_left is not None else "",
        "premium_inr": _fmt_decimal(r.premium_inr),
        "sum_insured_inr": _fmt_decimal(r.sum_insured_inr),
        "status": r.status,
    }


def _fmt_decimal(v: Any) -> str:
    if v is None:
        return ""
    return str(v)


def _fmt_date(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v)


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------

def build_xlsx(
    headers: list[str],
    column_keys: list[str],
    rows: list[dict[str, Any]],
) -> io.BytesIO:
    """Write rows to a new workbook and return a BytesIO buffer (position=0)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(k, "") for k in column_keys])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def write_policies_xlsx(
    policies: list[Any], provider_map: dict[uuid.UUID, str]
) -> io.BytesIO:
    headers = [label for _, label in EXPORT_COLUMNS]
    column_keys = [key for key, _ in EXPORT_COLUMNS]
    data_rows = [_policy_row(p, provider_map) for p in policies]
    return build_xlsx(headers, column_keys, data_rows)


def write_renewals_xlsx(rows: list[RenewalReportRow]) -> io.BytesIO:
    headers = [
        "Policy ID", "Title", "Category", "Provider",
        "Expiry Date", "Days Left", "Premium (INR)", "Sum Insured (INR)", "Status",
    ]
    column_keys = [
        "policy_id", "title", "category", "provider_name",
        "expiry_date", "days_left", "premium_inr", "sum_insured_inr", "status",
    ]
    data_rows = [_renewal_row(r) for r in rows]
    return build_xlsx(headers, column_keys, data_rows)


# ---------------------------------------------------------------------------
# CSV writer
# ---------------------------------------------------------------------------

def write_policies_csv(
    policies: list[Any], provider_map: dict[uuid.UUID, str]
) -> io.BytesIO:
    headers = [label for _, label in EXPORT_COLUMNS]
    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(wrapper, fieldnames=headers)
    writer.writeheader()
    for p in policies:
        row = _policy_row(p, provider_map)
        writer.writerow({label: row[key] for key, label in EXPORT_COLUMNS})
    wrapper.flush()
    wrapper.detach()
    buf.seek(0)
    return buf


def write_renewals_csv(rows: list[RenewalReportRow]) -> io.BytesIO:
    headers = [
        "Policy ID", "Title", "Category", "Provider",
        "Expiry Date", "Days Left", "Premium (INR)", "Sum Insured (INR)", "Status",
    ]
    buf = io.BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8-sig", newline="")
    writer = csv.DictWriter(wrapper, fieldnames=headers)
    writer.writeheader()
    for r in rows:
        rd = _renewal_row(r)
        writer.writerow({
            "Policy ID": rd["policy_id"],
            "Title": rd["title"],
            "Category": rd["category"],
            "Provider": rd["provider_name"],
            "Expiry Date": rd["expiry_date"],
            "Days Left": rd["days_left"],
            "Premium (INR)": rd["premium_inr"],
            "Sum Insured (INR)": rd["sum_insured_inr"],
            "Status": rd["status"],
        })
    wrapper.flush()
    wrapper.detach()
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Import parser
# ---------------------------------------------------------------------------

def _normalize_category(raw: str) -> str | None:
    """Map a freeform category string to a valid PolicyCategory or None."""
    cleaned = raw.strip().lower().replace(" ", "_").replace("-", "_")
    if cleaned in _VALID_CATEGORIES:
        return cleaned
    return None


def _parse_date_cell(value: Any) -> date | None:
    """Accept an Excel date object, a datetime, or YYYY-MM-DD string."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except InvalidOperation:
        return None


def _map_header(raw: str) -> str | None:
    """Map a raw column header (case-insensitive) to the canonical field name."""
    return _IMPORT_ALIASES.get(raw.strip().lower())


def parse_import_rows(
    raw_rows: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    """Normalize raw parsed rows to canonical field dicts.

    Each output dict has keys matching PolicyCreate fields plus 'provider_name'.
    Unknown / missing columns are silently ignored; required ones are validated
    by the caller.
    """
    result = []
    for row in raw_rows:
        mapped: dict[str, Any] = {}
        for col, val in row.items():
            field = _map_header(str(col))
            if field:
                mapped[field] = val
        result.append(mapped)
    return result


async def run_import(
    db: AsyncSession,
    user: CurrentUser,
    raw_rows: list[dict[str, Any]],
    default_org_id: uuid.UUID,
) -> tuple[int, int, list[ImportRowError]]:
    """Create policies from parsed rows.

    Returns (created_count, skipped_count, errors).
    Skipped = blank rows.  Errors = rows with validation problems (not aborted).
    """
    created = 0
    skipped = 0
    errors: list[ImportRowError] = []

    if len(raw_rows) > MAX_IMPORT_ROWS:
        raise AppError(
            ErrorCode.validation_error,
            f"Import file exceeds the {MAX_IMPORT_ROWS}-row limit "
            f"({len(raw_rows)} rows found). Split the file and retry.",
        )

    normalized = parse_import_rows(raw_rows)

    for idx, row in enumerate(normalized, start=2):  # row 1 = header
        # Blank row detection — all values are empty strings or None.
        if all((v is None or str(v).strip() == "") for v in row.values()):
            skipped += 1
            continue

        # Required: title
        title = str(row.get("title", "")).strip()
        if not title:
            errors.append(ImportRowError(row=idx, message="'Title' is required"))
            continue

        # Required: category
        raw_cat = str(row.get("category", "")).strip()
        category = _normalize_category(raw_cat)
        if not category:
            errors.append(
                ImportRowError(
                    row=idx,
                    message=f"Invalid category '{raw_cat}'. "
                    f"Must be one of: {', '.join(sorted(_VALID_CATEGORIES))}",
                )
            )
            continue

        payload = PolicyCreate(
            org_id=default_org_id,
            category=category,  # type: ignore[arg-type]
            title=title,
            policy_number=str(row["policy_number"]).strip() or None
            if row.get("policy_number")
            else None,
            sum_insured_inr=_parse_decimal(row.get("sum_insured_inr")),
            premium_inr=_parse_decimal(row.get("premium_inr")),
            gst_inr=_parse_decimal(row.get("gst_inr")),
            inception_date=_parse_date_cell(row.get("inception_date")),
            expiry_date=_parse_date_cell(row.get("expiry_date")),
            renewal_date=_parse_date_cell(row.get("renewal_date")),
        )

        try:
            await create_policy(db, user, payload)
            created += 1
        except Exception as exc:  # noqa: BLE001
            errors.append(ImportRowError(row=idx, message=str(exc)))

    return created, skipped, errors


# ---------------------------------------------------------------------------
# File parsers — convert raw bytes to list[dict[str, Any]]
# ---------------------------------------------------------------------------

def parse_csv_bytes(data: bytes) -> list[dict[str, Any]]:
    """Parse CSV bytes (UTF-8 or UTF-8-BOM) into a list of row dicts."""
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def parse_xlsx_bytes(data: bytes) -> list[dict[str, Any]]:
    """Parse XLSX bytes into a list of row dicts (first row = headers)."""
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    result = []
    for row in rows[1:]:
        result.append(dict(zip(headers, row, strict=False)))
    return result
